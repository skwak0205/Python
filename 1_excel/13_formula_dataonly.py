from openpyxl import load_workbook

# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active
#
# # 수식 그대로 가져옴
# for row in ws.values:
#     for cell in row:
#         print(cell)

wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# 수식이 아닌 실제 데이터 가져옴
# evaluate 되지 않은 상태의 데이터는 None으로 가져옴 (파일 한번 저장 후 실행하면 데이터 가져옴)
for row in ws.values:
    for cell in row:
        print(cell)
