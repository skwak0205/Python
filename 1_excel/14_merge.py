from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 병합하기
# ws.merge_cells("B2:D2") # B2부터 D2까지 합침
# ws["B2"].value = "Merged Cell"

# 병합 해제
ws.unmerge_cells("B2:D2")

wb.save("sample_merge.xlsx")
