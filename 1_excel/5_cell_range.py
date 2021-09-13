from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 한줄씩 data 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11):  # 10개의 row에 데이터 입력
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"]  # 영어 column 만 가져오기
for cell in col_B:
    print(cell.value)

col_range = ws["B:C"]  # 영어, 수학 column 함께 가져오기
for cols in col_range:
    for cell in cols:
        print(cell.value)

row_title = ws[1]  # 첫번째 row만 가져오기
for cell in row_title:
    print(cell.value)

row_range = ws[2:6]  # 2번째에서 6번째 줄까지 가져오기
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

from openpyxl.utils.cell import coordinate_from_string

row_range2 = ws[2:ws.max_row]  # 2번째에서 마지막 줄까지
for rows in row_range2:
    for cell in rows:
        print(cell.value, end=" ")
        print(cell.coordinate, end=" ")  # 각 cell에 대한 좌표 정보 가져옴
        xy = coordinate_from_string(cell.coordinate)  # 문자와 숫자 나눠서 출력
        print(xy, end=" ")
    print()

# 전체 rows
print(tuple(ws.rows))
for row in tuple(ws.rows):
    print(row[1].value)

# 전체 columns
print(tuple(ws.columns))
for column in tuple(ws.columns):
    print(column[0].value)

for row in ws.iterrows():  # 전체 row
    print(row)

# 1번째 줄부터 5번째 줄까지, 2번째 열부터 3번째 열까지
# 파라미터는 생략 가능
for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):
    print(row)

for col in ws.iter_cols():  # 전체 column
    print(col[0].value)

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)

wb.save("sample.xlsx")
