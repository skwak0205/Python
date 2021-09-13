from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet()  # 새로운 sheet 기본 이름으로 생성
ws.title = "MySheet"  # Sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff"  # RGB 형태로 Tab 색을 변경

ws1 = wb.create_sheet("YourSheet")  # 주어진 이름으로 새 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2)  # 인덱스로 넣고 싶은 자리에 Sheet 생성

new_ws = wb["NewSheet"]  # Dict 형태로 Sheet에 접근 가능

print(wb.sheetnames)  # 모든 sheet들의 이름 출력

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")
